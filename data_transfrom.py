from openpyxl import load_workbook

workbook = load_workbook(filename='data.xlsx')

sheet = workbook.active

for row in sheet.iter_rows(values_only=True):
    print(row)

workbook.close()
